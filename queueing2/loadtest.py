import multiprocessing
from multiprocessing import Manager
import requests
import random
import json
import time
from openpyxl import load_workbook

def run_instance(instance_number, url, headers, arr):
    print(f"Instance {instance_number} is running")

    body = {
        "prompt": "some text"
    }

    start_time = time.time()
    try:
        response = requests.post(url, headers=headers, data=json.dumps(body))
        response.raise_for_status()
        response_data = response.json()
        elapsed_time = time.time() - start_time

        arr[instance_number][0] = instance_number
        arr[instance_number][1] = start_time
        arr[instance_number][2] = elapsed_time
        arr[instance_number][3] = response_data.get("response", "")

        print(f"Process: {instance_number} | Elapsed Time: {elapsed_time}")
        print(response_data.get("response", ""))
        print("")
    except requests.exceptions.RequestException as e:
        elapsed_time = time.time() - start_time
        arr[instance_number][0] = instance_number
        arr[instance_number][1] = start_time
        arr[instance_number][2] = elapsed_time
        arr[instance_number][3] = f"Request failed: {e}"

        print(f"Status Code : {response.status_code if response else 'No response'}")
        print(f"Request failed: {e}")

if __name__ == "__main__":
    num_instances = 100  # Number of instances
    processes = []

    url = "http://localhost:8000/process"

    headers = {
        "Content-Type": "application/json"
    }

    manager = Manager()

    # Create a 2D list for shared data
    arr = manager.list([manager.list([None] * 4) for _ in range(num_instances)])

    for i in range(num_instances):
        p = multiprocessing.Process(target=run_instance, args=(i, url, headers, arr))
        processes.append(p)
        p.start()

    for p in processes:
        p.join()

    xlsx_file = "/home/tisuper/Desktop/python/queueing2/queue.xlsx"
    wb = load_workbook(xlsx_file)
    ws = wb.active

    for i in range(num_instances):
        for j in range(4):  # We have 4 columns now
            ws.cell(row=i + 1, column=j + 1, value=arr[i][j])

    wb.save(xlsx_file)

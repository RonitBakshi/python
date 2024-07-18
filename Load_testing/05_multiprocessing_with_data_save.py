import multiprocessing
from multiprocessing import Manager
import requests
import random
import json
import time
from openpyxl import load_workbook

def run_instance(instance_number, url, prompts, character_names, headers, num_req, arr):
    print(f"Instance {instance_number} is running")

    for i in range(num_req):
        prompt = random.choice(prompts).strip()
        character = random.choice(character_names)

        body = {
            "messages": [
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            "mode": "chat",
            "model": "chronos-hermes-13b-v2.Q4_K_M.gguf",
            "character": character,
            "name1": "Sumit",
            "name2": character,
            "stream": False,
            "temperature": 2,
            "top_p": 0.7,
            "top_k": 100,
            "repetition_penalty": 1,
            "max_tokens": 4096,
            "user": "Sumit",
            "guidance_scale": 1
        }

        start_time = time.time()
        response = requests.post(url, headers=headers, data=json.dumps(body))
        end_time = time.time()

        elapsed_time = end_time - start_time

        if response.status_code == 200:
            arr[i][instance_number] = elapsed_time
            print(f"Process: {instance_number} | Generation: {i} | Time:{elapsed_time}")
        else:
            print(f"Status Code : {response.status_code}")
            arr[i][instance_number] = f"Status code: {response.status_code} : {elapsed_time}"


if __name__ == "__main__":
    num_instances = 1000   # Number of instances
    num_req =  100        # Number of requests per instance
    processes = []

    url = "http://192.168.1.34:5000/v1/chat/completions"

    with open("Load_testing/assets/prompts.txt", "r") as file:
        prompts = file.readlines()

    character_names = [
        "Marlowe",
        "Cassy",
        "Example"
    ]

    headers = {
        "Content-Type": "application/json"
    }

    manager = Manager()

    # Create a 2D list for shared data
    arr = manager.list([manager.list([None] * num_instances) for _ in range(num_req)])

    for i in range(num_instances):
        p = multiprocessing.Process(target=run_instance, args=(i, url, prompts, character_names, headers, num_req, arr))
        processes.append(p)
        p.start()

    for p in processes:
        p.join()

    xlsx_file = "/home/tisuper/Desktop/python/Load_testing/assets/data.xlsx"
    wb = load_workbook(xlsx_file)
    ws = wb.active

    for i in range(num_req):
        for j in range(num_instances):
            ws.cell(row=i + 1, column=j + 1, value=arr[i][j])

    wb.save(xlsx_file)

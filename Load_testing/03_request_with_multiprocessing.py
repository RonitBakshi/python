import multiprocessing
import requests
import random
import json
import time
from openpyxl import load_workbook

def run_instance(instance_number,url,prompts,character_names,headers,ws):
   
    print(f"Instance {instance_number} is running")
    
    for i in range(3):
     
        prompt = random.choice(prompts).strip() 
        character = random.choice(character_names)

        body = {
            "messages": [
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            "mode": "chat",
            "model": "chronos-hermes-13b-v2.Q4_K_M.gguf",
            "character": character,
            "name1": "Sumit",
            "name2": character,
            "stream": False,
            "temperature": 2,
            "top_p": 0.7,
            "top_k": 100,
            "repetition_penalty": 1,
            "max_tokens": 4096,
            "user": "Sumit",
            "guidance_scale": 1
        }

        start_time = time.time()
        response = requests.post(url, headers=headers, data=json.dumps(body))
        end_time = time.time()

        if response.status_code == 200:
            print(end_time-start_time)
            ws.cell(row=i+1, column=instance_number+1, value=end_time-start_time)
           
        else:
            print(f"Status Code : {response.status_code}")
        


if __name__ == "__main__":
    num_instances = 10
    processes = []

    url = "http://192.168.1.34:5000/v1/chat/completions"

    with open("Load_testing/assets/prompts.txt", "r") as file:
        prompts = file.readlines()

    character_names = [
        "Marlowe",
        "Cassy",
        "Example"
    ]

    headers = {
        "Content-Type": "application/json"
    }

    xlsx_file = "/home/tisuper/Desktop/python/Load_testing/assets/data.xlsx"
    wb = load_workbook(xlsx_file)
    ws = wb.active 

    for i in range(num_instances):
        p = multiprocessing.Process(target=run_instance, args=(i,url,prompts,character_names,headers,ws))
        processes.append(p)
        p.start()

    for p in processes:
        p.join()

    wb.save(xlsx_file)